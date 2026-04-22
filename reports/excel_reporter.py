"""
Excel 보고서 생성기
5개 시트: 요약, 사용자별 현황, 개인정보 검출 상세, 과다조회 상세, 소명 요청 양식
"""
from __future__ import annotations
import os
from datetime import datetime, date
from pathlib import Path
from models.user_summary import UserSummary
from models.finding import Finding

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
        GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

RISK_KR = {'CRITICAL': '최고위험', 'HIGH': '고위험', 'MEDIUM': '중위험', 'LOW': '저위험'}

# ── 색상 정의 ─────────────────────────────────────────────
COLORS = {
    'CRITICAL': 'FFCCCC',   # 연빨강
    'HIGH': 'FFE5CC',       # 연주황
    'MEDIUM': 'FFFACC',     # 연노랑
    'LOW': 'FFFFFF',        # 흰색
    'HEADER': '1F4E79',     # 진파랑 (헤더)
    'SUBHEADER': '2E75B6',  # 중간파랑
    'LABEL': 'D6E4F0',      # 연파랑 (레이블)
    'BORDER': '8EA9C1',     # 테두리
}


def _make_fill(hex_color: str) -> 'PatternFill':
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _make_border() -> 'Border':
    thin = Side(style='thin', color=COLORS['BORDER'])
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_font(white: bool = True) -> 'Font':
    return Font(bold=True, color='FFFFFF' if white else '1F4E79', name='맑은 고딕', size=10)


def _normal_font() -> 'Font':
    return Font(name='맑은 고딕', size=9)


def _center() -> 'Alignment':
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def _left() -> 'Alignment':
    return Alignment(horizontal='left', vertical='center', wrap_text=True)


def auto_fit_columns(ws, min_width: int = 8, max_width: int = 50) -> None:
    """열 너비를 내용에 맞게 자동 조정합니다."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ''
                # 한국어는 영문의 약 2배 너비
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, char_len)
            except Exception:
                pass
        adjusted = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col_letter].width = adjusted


def generate_excel(
    summaries: list[UserSummary],
    start_date: date,
    end_date: date,
    output_path: str,
    total_events: int = 0,
    total_lines: int = 0,
    deltas_week: dict | None = None,
    deltas_month: dict | None = None,
    week_period: str = '',
    month_period: str = '',
) -> str:
    """Excel 보고서를 생성하고 파일 경로를 반환합니다."""
    if not OPENPYXL_AVAILABLE:
        print("[경고] openpyxl이 설치되지 않아 Excel 보고서를 생성할 수 없습니다.")
        return ''

    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    report_ts = datetime.now()

    # 소명 우선순위 생성
    from reports.justification_builder import build_justification_list
    just_items = build_justification_list(summaries)

    # 시트 생성
    _create_summary_sheet(wb, summaries, start_date, end_date, report_ts, total_events, total_lines)
    _create_user_overview_sheet(wb, summaries)
    _create_pii_detail_sheet(wb, summaries)
    _create_excess_detail_sheet(wb, summaries)
    _create_justification_sheet(wb, just_items, start_date, end_date)
    _create_evidence_detail_sheet(wb, just_items)

    # 비교 분석 시트 (이력 데이터 있을 때만) — Sheet 7
    if deltas_week or deltas_month:
        _create_comparison_sheet(
            wb, summaries,
            deltas_week or {}, deltas_month or {},
            week_period, month_period,
            start_date, end_date,
        )

    # 저장
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _create_summary_sheet(wb, summaries, start_date, end_date, report_ts, total_events, total_lines):
    ws = wb.create_sheet("1. 요약")
    ws.sheet_view.showGridLines = False

    header_fill = _make_fill(COLORS['HEADER'])
    label_fill = _make_fill(COLORS['LABEL'])
    border = _make_border()

    def write_section_header(row, title):
        cell = ws.cell(row=row, column=1, value=title)
        cell.fill = _make_fill(COLORS['SUBHEADER'])
        cell.font = _header_font()
        cell.alignment = _left()
        ws.merge_cells(f'A{row}:C{row}')

    def write_row(row, label, value, value_fill=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = label_fill
        lc.font = Font(bold=True, name='맑은 고딕', size=9)
        lc.border = border

        vc = ws.cell(row=row, column=2, value=value)
        vc.font = _normal_font()
        vc.border = border
        if value_fill:
            vc.fill = _make_fill(value_fill)
        ws.merge_cells(f'B{row}:C{row}')

    # 제목
    ws.merge_cells('A1:C1')
    title_cell = ws['A1']
    title_cell.value = '개인정보 오남용·과다조회 점검 보고서'
    title_cell.fill = _make_fill(COLORS['HEADER'])
    title_cell.font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=14)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 35

    ws.row_dimensions[2].height = 10

    write_section_header(3, '■ 분석 정보')
    write_row(4, '분석 기간', f"{start_date} ~ {end_date}")
    write_row(5, '보고서 생성 일시', report_ts.strftime('%Y-%m-%d %H:%M:%S'))
    write_row(6, '총 분석 로그 라인 수', f"{total_lines:,} 줄")
    write_row(7, '총 분석 이벤트 수', f"{total_events:,} 건")

    ws.row_dimensions[8].height = 10

    # 통계 계산
    total_users = len(summaries)
    critical_users = sum(1 for s in summaries if s.risk_level == 'CRITICAL')
    high_users = sum(1 for s in summaries if s.risk_level == 'HIGH')
    medium_users = sum(1 for s in summaries if s.risk_level == 'MEDIUM')
    total_findings = sum(len(s.findings) for s in summaries)
    pii_findings = sum(s.pii_finding_count for s in summaries)
    excess_findings = sum(s.excess_finding_count for s in summaries)
    total_exposed = sum(s.total_pii_records_exposed for s in summaries)
    max_single = max((s.max_single_query_exposure for s in summaries), default=0)

    all_pii_types: set = set()
    for s in summaries:
        all_pii_types.update(s.pii_types_seen)

    write_section_header(9, '■ 점검 결과 요약')
    write_row(10, '분석 대상 사용자 수', f"{total_users:,} 명")
    write_row(11, '최고위험(CRITICAL) 사용자', f"{critical_users:,} 명", 'FFCCCC' if critical_users > 0 else None)
    write_row(12, '고위험(HIGH) 사용자', f"{high_users:,} 명", 'FFE5CC' if high_users > 0 else None)
    write_row(13, '중위험(MEDIUM) 사용자', f"{medium_users:,} 명", 'FFFACC' if medium_users > 0 else None)
    write_row(14, '총 이상 징후 건수', f"{total_findings:,} 건")
    write_row(15, '개인정보 노출 이상 건수', f"{pii_findings:,} 건")
    write_row(16, '과다조회 이상 건수', f"{excess_findings:,} 건")
    write_row(17, '검출된 개인정보 유형', ', '.join(sorted(all_pii_types)) if all_pii_types else '없음')

    ws.row_dimensions[18].height = 10
    write_section_header(19, '■ 실효 개인정보 노출량 요약')
    write_row(20, '전체 PII 노출 레코드 수 (합계)', f"{total_exposed:,} 건",
              'FFCCCC' if total_exposed >= 50_000 else ('FFE5CC' if total_exposed >= 10_000 else None))
    write_row(21, '단일 쿼리 최대 노출 건수', f"{max_single:,} 건",
              'FFCCCC' if max_single >= 20_000 else ('FFE5CC' if max_single >= 5_000 else None))
    select_pii_users = sum(1 for s in summaries if s.select_pii_query_count > 0)
    write_row(22, 'SELECT절 PII 조회 발생 사용자 수', f"{select_pii_users:,} 명")

    ws.row_dimensions[23].height = 10
    write_section_header(24, '■ 위험 등급 기준')
    ws.cell(row=25, column=1, value='등급').font = Font(bold=True, name='맑은 고딕', size=9)
    ws.cell(row=25, column=2, value='점수 범위').font = Font(bold=True, name='맑은 고딕', size=9)
    ws.cell(row=25, column=3, value='조치 기준').font = Font(bold=True, name='맑은 고딕', size=9)

    levels = [
        ('CRITICAL (최고위험)', '70-100점', '즉각 조사 및 소명 요청', 'FFCCCC'),
        ('HIGH', '45-69점', '소명 요청 및 모니터링 강화', 'FFE5CC'),
        ('MEDIUM', '20-44점', '주의 조치 및 교육', 'FFFACC'),
        ('LOW', '0-19점', '정상 범위', 'FFFFFF'),
    ]
    for i, (lvl, score_range, action, color) in enumerate(levels, start=26):
        ws.cell(row=i, column=1, value=lvl).fill = _make_fill(color)
        ws.cell(row=i, column=2, value=score_range)
        ws.cell(row=i, column=3, value=action)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35


def _create_user_overview_sheet(wb, summaries: list[UserSummary]):
    ws = wb.create_sheet("2. 사용자별 현황")
    ws.sheet_view.showGridLines = False

    headers = [
        '사원ID', '총조회건수', 'PII쿼리수', 'PII유형',
        'PII출력건수', '단일쿼리최대출력', 'PII컬럼조회수',
        '최대시간당조회', '최대일당조회', '피크시간', '피크날짜',
        '야간조회수', '대량조회수', '이상건수', '위험점수', '위험등급'
    ]

    header_fill = _make_fill(COLORS['HEADER'])
    border = _make_border()

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = border

    ws.row_dimensions[1].height = 22

    for row, summary in enumerate(summaries, start=2):
        level_color = COLORS.get(summary.risk_level, 'FFFFFF')
        values = [
            summary.user_id,
            summary.total_events,
            summary.pii_event_count,
            summary.pii_types_str,
            summary.total_pii_records_exposed,
            summary.max_single_query_exposure,
            summary.select_pii_query_count,
            summary.max_queries_per_hour,
            summary.max_queries_per_day,
            summary.peak_hour,
            summary.peak_day,
            summary.after_hours_count,
            summary.bulk_export_count,
            summary.flagged_event_count,
            summary.risk_score,
            RISK_KR.get(summary.risk_level, summary.risk_level),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = _make_fill(level_color)
            cell.font = _normal_font()
            cell.alignment = _center() if col != 4 else _left()
            cell.border = border

    auto_fit_columns(ws)
    ws.freeze_panes = 'A2'


def _create_pii_detail_sheet(wb, summaries: list[UserSummary]):
    ws = wb.create_sheet("3. 개인정보 검출 상세")
    ws.sheet_view.showGridLines = False

    headers = ['사원ID', '일시', 'PII유형', '검출원본값', '노출유형', '노출레코드수', '위험등급', '증적(컨텍스트)', '소스파일:라인']
    header_fill = _make_fill(COLORS['HEADER'])
    border = _make_border()

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = border

    ws.row_dimensions[1].height = 22
    row = 2
    MAX_ROWS = 100_000

    for summary in summaries:
        for finding in summary.findings:
            if finding.category != 'PII_EXPOSURE':
                continue
            if row > MAX_ROWS:
                ws.cell(row=row, column=1, value=f'[이하 {len(summaries)} 건 이상 - 전체 결과는 원본 로그 참조]')
                break
            level_color = COLORS.get(summary.risk_level, 'FFFFFF')
            exposure_type = finding.details.get('exposure_type', '') if finding.details else ''
            result_rows = finding.details.get('result_row_count') if finding.details else None
            orig_val = finding.details.get('original_value', '') if finding.details else ''
            exposure_type_kr = {'FULL_EXPOSURE': '전체컬럼노출', 'PARTIAL_EXPOSURE': 'PII컬럼직접출력',
                                'SEARCH_ONLY': '검색조건만사용', 'NONE': '없음'}.get(exposure_type, exposure_type)
            values = [
                finding.user_id,
                finding.timestamp_str,
                finding.pii_types_str,
                orig_val,                          # 검출원본값 (비마스킹)
                exposure_type_kr,
                result_rows if result_rows is not None else '건수미확인',
                RISK_KR.get(summary.risk_level, summary.risk_level),
                finding.evidence[:300] if finding.evidence else '',
                finding.raw_reference,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _make_fill(level_color)
                cell.font = _normal_font()
                cell.alignment = _left()
                cell.border = border
            row += 1

    auto_fit_columns(ws)
    ws.freeze_panes = 'A2'


def _create_excess_detail_sheet(wb, summaries: list[UserSummary]):
    ws = wb.create_sheet("4. 과다조회 상세")
    ws.sheet_view.showGridLines = False

    headers = ['사원ID', '일시', '이상유형', '위험등급', '내용', '소스파일:라인']
    header_fill = _make_fill(COLORS['HEADER'])
    border = _make_border()

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = border

    ws.row_dimensions[1].height = 22
    row = 2

    category_names = {
        'PII_EXPOSURE':     'PII 오남용',
        'PII_RECORD_EXPOSURE': 'PII 레코드 노출',
        'EXCESSIVE_ACCESS': '과다조회',
        'AFTER_HOURS':      '업무외시간조회',
        'BULK_EXPORT':      '대량조회',
    }

    for summary in summaries:
        for finding in summary.findings:
            level_color = COLORS.get(summary.risk_level, 'FFFFFF')
            values = [
                finding.user_id,
                finding.timestamp_str,
                category_names.get(finding.category, finding.category),
                RISK_KR.get(summary.risk_level, summary.risk_level),
                finding.evidence[:300] if finding.evidence else '',
                finding.raw_reference,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _make_fill(level_color)
                cell.font = _normal_font()
                cell.alignment = _left()
                cell.border = border
            row += 1

    auto_fit_columns(ws)
    ws.freeze_panes = 'A2'


def _create_justification_sheet(wb, just_items: list, start_date, end_date):
    """소명 요청 양식 시트 - 우선순위 순 정렬, 상세 근거 포함"""
    ws = wb.create_sheet("5. 소명 요청 양식")
    ws.sheet_view.showGridLines = False
    border = _make_border()

    # ── 제목 ─────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    ws['A1'] = f'개인정보 오남용·과다조회 소명 요청 자료  |  분석 기간: {start_date} ~ {end_date}'
    ws['A1'].fill = _make_fill('8B0000')
    ws['A1'].font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=12)
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    ws['A2'] = ('※ 본 양식은 개인정보 오남용·과다조회 의심 사용자에 대한 소명 자료 징구용입니다. '
                '소명 내용은 당사자가 직접 기재하여 담당부서에 제출하시기 바랍니다.')
    ws['A2'].fill = _make_fill('FFE5CC')
    ws['A2'].font = Font(name='맑은 고딕', size=9, italic=True)
    ws['A2'].alignment = _left()
    ws.row_dimensions[2].height = 20

    # ── 헤더 ─────────────────────────────────────────────
    headers = [
        '소명\n우선순위', '사원ID', '위험\n등급', '위험\n점수',
        '주요 위반 이유 (소명 근거)',
        '소명 요청 질문',
        '소명 내용\n(당사자 기입)',
        '처리 결과\n/ 담당자'
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = _make_fill('8B0000')
        cell.font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=9)
        cell.alignment = _center()
        cell.border = border
    ws.row_dimensions[3].height = 30

    if not just_items:
        ws.merge_cells('A4:H4')
        ws['A4'] = '소명 요청 대상자 없음 (MEDIUM 이상 등급 사용자 없음)'
        ws['A4'].font = _normal_font()
    else:
        urgency_colors = {'즉시': 'FFCCCC', '긴급': 'FFE5CC', '검토': 'D6E4F0'}
        row = 4
        for item in just_items:
            level_color = COLORS.get(item.risk_level, 'FFFFFF')
            urg_color   = urgency_colors.get(item.urgency, 'FFFFFF')

            reasons_txt  = '\n'.join(f'• {r}' for r in item.reasons)
            questions_txt = '\n'.join(f'Q{i+1}. {q}' for i, q in enumerate(item.questions))

            row_h = max(80, len(item.reasons) * 18 + len(item.questions) * 20)
            ws.row_dimensions[row].height = row_h

            values_and_fills = [
                (f'#{item.priority_rank}\n{item.urgency}',  urg_color,   _center()),
                (item.user_id,                               level_color, _center()),
                (RISK_KR.get(item.risk_level, item.risk_level), level_color, _center()),
                (f'{item.risk_score:.1f}',                   level_color, _center()),
                (reasons_txt,                                level_color, _left()),
                (questions_txt,                              'FAFAFA',    _left()),
                ('',                                         'FAFAFA',    _left()),
                ('',                                         'FAFAFA',    _left()),
            ]
            for col, (val, fill_hex, align) in enumerate(values_and_fills, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _make_fill(fill_hex)
                cell.font = _normal_font()
                cell.alignment = align
                cell.border = border
            row += 1

    col_widths = [10, 13, 9, 8, 52, 48, 35, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'


def _create_evidence_detail_sheet(wb, just_items: list):
    """소명 증거 상세 시트 - 핵심 증거 로그 (마스킹), 화면 노출 추정"""
    ws = wb.create_sheet("6. 소명 증거 상세")
    ws.sheet_view.showGridLines = False
    border = _make_border()

    ws.merge_cells('A1:J1')
    ws['A1'] = '소명 증거 상세  |  노출량은 SELECT 쿼리 기반 추정값'
    ws['A1'].fill = _make_fill(COLORS['HEADER'])
    ws['A1'].font = _header_font()
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 22

    headers = [
        '사원ID', '우선순위', '위험등급',
        '발생일시', '위반유형', 'PII유형', '검출원본값',
        '노출유형', '반환건수', '증거(컨텍스트)', '참조'
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = _make_fill(COLORS['HEADER'])
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = border
    ws.row_dimensions[2].height = 22

    row = 3
    for item in just_items:
        # 사용자 구분 헤더 행 삽입
        ws.merge_cells(f'A{row}:K{row}')
        user_header = ws.cell(row=row, column=1,
                              value=f'▶ 사원ID: {item.user_id}  |  위험등급: {RISK_KR.get(item.risk_level, item.risk_level)}  |  위험점수: {item.risk_score:.1f}  |  소명우선순위: #{item.priority_rank}')
        user_header.fill = _make_fill('2E4057')
        user_header.font = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
        user_header.alignment = _left()
        user_header.border = border
        ws.row_dimensions[row].height = 20
        row += 1

        for f in item.key_findings:
            sev_color = COLORS.get(f.get('severity', 'LOW'), 'FFFFFF')
            values = [
                item.user_id,
                f'#{item.priority_rank}',
                f.get('severity_kr', ''),
                f.get('timestamp_str', ''),
                f.get('category_kr', ''),
                f.get('pii_types_str', ''),
                f.get('original_value', ''),        # 검출원본값 (비마스킹)
                f.get('exposure_type_kr', ''),
                f.get('result_rows') if f.get('result_rows') is not None else '미상',
                (f.get('evidence') or '')[:300],
                f.get('raw_reference', ''),
            ]
            ws.row_dimensions[row].height = 35
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _make_fill(sev_color)
                cell.font = _normal_font()
                cell.alignment = _left()
                cell.border = border
                # 사원ID 컬럼 굵게 강조
                if col == 1:
                    cell.font = Font(bold=True, name='맑은 고딕', size=9)
            row += 1

        # 화면 노출 추정 섹션 (구분)
        ws.merge_cells(f'A{row}:K{row}')
        est_cell = ws.cell(row=row, column=1,
                           value=f'[{item.user_id}] 화면 노출 추정: {item.screen_estimate[:300]}')
        est_cell.fill = _make_fill('D6E4F0')
        est_cell.font = Font(name='맑은 고딕', size=8, italic=True)
        est_cell.alignment = _left()
        est_cell.border = border
        ws.row_dimensions[row].height = 45
        row += 1

    col_widths = [16, 8, 10, 18, 16, 18, 22, 14, 10, 60, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'B3'


def _create_comparison_sheet(
    wb,
    summaries: list[UserSummary],
    deltas_week: dict,
    deltas_month: dict,
    week_period: str,
    month_period: str,
    start_date,
    end_date,
):
    """6. 비교 분석 시트 - 이전 주/이전 달과의 지표 변화를 표시합니다."""
    ws = wb.create_sheet("7. 비교 분석")
    ws.sheet_view.showGridLines = False
    border = _make_border()

    # ── 안내 행 ──────────────────────────────────────────────
    ws.merge_cells('A1:P1')
    ws['A1'] = f'비교 분석 | 현재 기간: {start_date} ~ {end_date}'
    ws['A1'].fill = _make_fill(COLORS['HEADER'])
    ws['A1'].font = _header_font()
    ws['A1'].alignment = _center()
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:P2')
    ws['A2'] = (
        f'1주 전 비교 기간: {week_period or "이력 없음"}  |  '
        f'1개월 전 비교 기간: {month_period or "이력 없음"}  |  '
        f'[↑] 증가(위험 상승)  [↓] 감소(위험 감소)  [→] 변화없음  [-] 이력 없음'
    )
    ws['A2'].fill = _make_fill(COLORS['LABEL'])
    ws['A2'].font = Font(name='맑은 고딕', size=9, italic=True)
    ws['A2'].alignment = _left()
    ws.row_dimensions[2].height = 18

    # ── 헤더 ─────────────────────────────────────────────────
    headers = [
        '사원ID', '현재\n위험등급', '현재\n위험점수',
        '1주전\n위험점수', '점수변화\n(주)', '점수추세\n(주)',
        '1개월전\n위험점수', '점수변화\n(월)', '점수추세\n(월)',
        '현재\nPII건수', 'PII변화\n(주)', 'PII변화\n(월)',
        '현재\n일최대조회', '일최대변화\n(주)', '일최대변화\n(월)',
        '현재\n야간조회',
    ]
    header_fill = _make_fill(COLORS['HEADER'])
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = border
    ws.row_dimensions[3].height = 30

    def delta_cell(ws, row, col, delta_val, is_risk_field=True):
        """
        변화량 셀을 작성합니다.
        위험 필드: 증가=빨강, 감소=초록
        """
        if delta_val is None:
            cell = ws.cell(row=row, column=col, value='-')
            cell.font = _normal_font()
            cell.alignment = _center()
            cell.border = border
            return

        if isinstance(delta_val, float):
            display = f"{delta_val:+.1f}"
        else:
            display = f"{delta_val:+d}"

        cell = ws.cell(row=row, column=col, value=display)
        cell.font = _normal_font()
        cell.alignment = _center()
        cell.border = border

        if is_risk_field:
            if delta_val > 0:
                cell.fill = _make_fill('FFCCCC')  # 빨강: 위험 증가
            elif delta_val < 0:
                cell.fill = _make_fill('CCFFCC')  # 초록: 위험 감소
        else:
            if delta_val != 0:
                cell.fill = _make_fill('E8F4FD')  # 연파랑: 변화 있음

    def arrow_cell(ws, row, col, delta_val):
        """추세 화살표 셀"""
        if delta_val is None:
            val = '-'
        elif delta_val > 0:
            val = '↑ 증가'
        elif delta_val < 0:
            val = '↓ 감소'
        else:
            val = '→ 유지'

        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _normal_font()
        cell.alignment = _center()
        cell.border = border
        if delta_val is not None:
            if delta_val > 0:
                cell.fill = _make_fill('FFCCCC')
            elif delta_val < 0:
                cell.fill = _make_fill('CCFFCC')

    # ── 데이터 행 ────────────────────────────────────────────
    for data_row, s in enumerate(summaries, start=4):
        dw = deltas_week.get(s.user_id, {})
        dm = deltas_month.get(s.user_id, {})

        level_color = COLORS.get(s.risk_level, 'FFFFFF')

        def prev_score(delta_dict, field, current_val):
            d = delta_dict.get(field)
            if d is None:
                return None
            return round(current_val - d, 2)

        # 기본 정보
        def base_cell(col, val):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.fill = _make_fill(level_color)
            cell.font = _normal_font()
            cell.alignment = _center()
            cell.border = border

        base_cell(1, s.user_id)
        base_cell(2, RISK_KR.get(s.risk_level, s.risk_level))
        base_cell(3, s.risk_score)

        # 1주 전 점수
        ws.cell(row=data_row, column=4,
                value=prev_score(dw, 'risk_score', s.risk_score)).border = border
        ws.cell(row=data_row, column=4).font = _normal_font()
        ws.cell(row=data_row, column=4).alignment = _center()

        delta_cell(ws, data_row, 5, dw.get('risk_score'), is_risk_field=True)
        arrow_cell(ws, data_row, 6, dw.get('risk_score'))

        # 1개월 전 점수
        ws.cell(row=data_row, column=7,
                value=prev_score(dm, 'risk_score', s.risk_score)).border = border
        ws.cell(row=data_row, column=7).font = _normal_font()
        ws.cell(row=data_row, column=7).alignment = _center()

        delta_cell(ws, data_row, 8, dm.get('risk_score'), is_risk_field=True)
        arrow_cell(ws, data_row, 9, dm.get('risk_score'))

        # PII 건수 변화
        base_cell(10, s.pii_event_count)
        delta_cell(ws, data_row, 11, dw.get('pii_event_count'), is_risk_field=True)
        delta_cell(ws, data_row, 12, dm.get('pii_event_count'), is_risk_field=True)

        # 일 최대 조회 변화
        base_cell(13, s.max_queries_per_day)
        delta_cell(ws, data_row, 14, dw.get('max_queries_per_day'), is_risk_field=True)
        delta_cell(ws, data_row, 15, dm.get('max_queries_per_day'), is_risk_field=True)

        # 야간 조회
        base_cell(16, s.after_hours_count)

    # ── 열 너비 ──────────────────────────────────────────────
    col_widths = [14, 10, 10, 10, 10, 10, 12, 10, 10, 10, 10, 10, 12, 12, 12, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'

    # ── 신규 위험 사용자 하이라이트 섹션 ─────────────────────
    # 이전엔 LOW였다가 HIGH/CRITICAL로 상승한 사용자 강조
    new_risk_users = []
    for s in summaries:
        dw = deltas_week.get(s.user_id, {})
        dm = deltas_month.get(s.user_id, {})
        prev_level_w = dw.get('risk_level_prev', '')
        prev_level_m = dm.get('risk_level_prev', '')
        if s.risk_level in ('HIGH', 'CRITICAL'):
            if prev_level_w in ('LOW', 'MEDIUM') or prev_level_m in ('LOW', 'MEDIUM'):
                new_risk_users.append((s, prev_level_w or prev_level_m))

    if new_risk_users:
        note_row = len(summaries) + 5
        ws.merge_cells(f'A{note_row}:P{note_row}')
        ws.cell(row=note_row, column=1,
                value='⚠ 신규 위험 상승 사용자 (이전 기간 대비 LOW/MEDIUM → HIGH/CRITICAL 상승)').fill = _make_fill('FFE5CC')
        ws.cell(row=note_row, column=1).font = Font(bold=True, name='맑은 고딕', size=9)

        for i, (s, prev_level) in enumerate(new_risk_users):
            r = note_row + 1 + i
            ws.cell(row=r, column=1, value=s.user_id).font = Font(bold=True, name='맑은 고딕', size=9)
            ws.cell(row=r, column=2, value=f"{prev_level} → {s.risk_level}").fill = _make_fill('FFCCCC')
            ws.cell(row=r, column=3, value=s.risk_score)
            ws.cell(row=r, column=4, value='즉시 소명 요청 필요').font = Font(color='FF0000', name='맑은 고딕', size=9)
